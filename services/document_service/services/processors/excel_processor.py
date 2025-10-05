"""
Copyright (c) 2025 LALO AI SYSTEMS, LLC. All rights reserved.

PROPRIETARY AND CONFIDENTIAL

This file is part of LALO AI Platform and is protected by copyright law.
Unauthorized copying, modification, distribution, or use of this software,
via any medium, is strictly prohibited without the express written permission
of LALO AI SYSTEMS, LLC.
"""

from typing import Dict, Any
import openpyxl
from models import Document, ProcessingResult

class ExcelProcessor:
    async def process(self, document: Document) -> ProcessingResult:
        try:
            # Load workbook from bytes
            workbook = openpyxl.load_workbook(document.content)
            
            # Extract metadata
            metadata: Dict[str, Any] = {
                "sheets": [],
                "properties": {}
            }
            
            # Get sheet information
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                sheet_info = {
                    "name": sheet,
                    "row_count": ws.max_row,
                    "column_count": ws.max_column,
                    "has_charts": len(ws._charts) > 0,
                    "has_tables": len(ws.tables) > 0
                }
                metadata["sheets"].append(sheet_info)
            
            # Get workbook properties
            if workbook.properties:
                metadata["properties"] = {
                    "title": workbook.properties.title,
                    "subject": workbook.properties.subject,
                    "creator": workbook.properties.creator,
                    "keywords": workbook.properties.keywords,
                    "category": workbook.properties.category
                }
            
            # Update document metadata
            document.metadata.update(metadata)
            
            return ProcessingResult(
                success=True,
                document_id=document.id,
                message="Excel document processed successfully"
            )
            
        except Exception as e:
            return ProcessingResult(
                success=False,
                document_id=document.id,
                errors=[f"Excel processing error: {str(e)}"]
            )
